import pandas as pd
from openpyxl import load_workbook

from src.reporting.excel_report import generate_excel_report


def test_generate_excel_report_creates_expected_sheets(tmp_path):
    dataframe = pd.DataFrame(
        [
            {
                "execution_timestamp": "2026-03-30T16:00:00",
                "page_heading": "Secure Area",
                "flash_message": "You logged into a secure area!",
                "current_url": "https://the-internet.herokuapp.com/secure",
                "login_success": True,
                "message_length": 30,
                "page_category": "secure_area",
            }
        ]
    )

    output_path = tmp_path / "test_report.xlsx"
    generate_excel_report(dataframe, output_path)

    workbook = load_workbook(output_path)

    assert "raw_data" in workbook.sheetnames
    assert "summary" in workbook.sheetnames
